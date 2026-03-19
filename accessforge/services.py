from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.text import slugify

from .models import DataField, DataRecord, DataTable


QUERY_OPERATOR_CHOICES = [
    ("equals", "等於"),
    ("not_equals", "不等於"),
    ("contains", "包含"),
    ("starts_with", "開頭是"),
    ("ends_with", "結尾是"),
    ("gt", "大於"),
    ("gte", "大於或等於"),
    ("lt", "小於"),
    ("lte", "小於或等於"),
    ("is_empty", "為空"),
    ("not_empty", "不為空"),
    ("is_true", "為是"),
    ("is_false", "為否"),
]


@dataclass(frozen=True)
class FieldPathSpec:
    token: str
    label: str
    field: DataField
    relation_field: DataField | None = None

    @property
    def is_joined(self) -> bool:
        return self.relation_field is not None


def get_visible_fields(table: DataTable, user) -> list[DataField]:
    if user is None:
        return list(table.ordered_fields)
    return [field for field in table.ordered_fields if field.can_view(user)]


def get_accessible_field_paths(
    table: DataTable,
    user,
    *,
    include_joined: bool = True,
) -> list[FieldPathSpec]:
    paths: list[FieldPathSpec] = []
    for field in get_visible_fields(table, user):
        paths.append(FieldPathSpec(token=field.slug, label=field.name, field=field))
        if not include_joined:
            continue
        if not field.is_relation or field.related_table is None:
            continue
        if user is not None and not field.related_table.has_role(user):
            continue
        for related_field in get_visible_fields(field.related_table, user):
            paths.append(
                FieldPathSpec(
                    token=f"{field.slug}__{related_field.slug}",
                    label=f"{field.name} -> {related_field.name}",
                    field=related_field,
                    relation_field=field,
                )
            )
    return paths


def resolve_field_path(
    table: DataTable,
    user,
    token: str,
    *,
    include_joined: bool = True,
) -> FieldPathSpec:
    available = {
        spec.token: spec
        for spec in get_accessible_field_paths(table, user, include_joined=include_joined)
    }
    try:
        return available[token]
    except KeyError as exc:
        raise ValidationError(f"欄位路徑「{token}」目前不可用。") from exc


def get_value_for_field_path(record: DataRecord, field_path: FieldPathSpec):
    if not field_path.is_joined:
        return record.data.get(field_path.field.slug)
    relation_value = record.data.get(field_path.relation_field.slug)
    if relation_value in (None, "") or field_path.relation_field.related_table is None:
        return None
    related_record = field_path.relation_field.related_table.records.filter(
        pk=int(relation_value)
    ).first()
    if related_record is None:
        return None
    return related_record.data.get(field_path.field.slug)


def display_value_for_field_path(record: DataRecord, field_path: FieldPathSpec) -> str:
    if not field_path.is_joined:
        return field_path.field.display_value(record.data.get(field_path.field.slug))
    relation_value = record.data.get(field_path.relation_field.slug)
    if relation_value in (None, "") or field_path.relation_field.related_table is None:
        return "-"
    related_record = field_path.relation_field.related_table.records.filter(
        pk=int(relation_value)
    ).first()
    if related_record is None:
        return "-"
    return field_path.field.display_value(related_record.data.get(field_path.field.slug))


def normalize_header(header: object, index: int) -> tuple[str, str]:
    label = str(header).strip() if header not in (None, "") else f"欄位 {index}"
    slug = slugify(label) or f"column_{index}"
    return label, slug


def parse_uploaded_rows(uploaded_file) -> tuple[list[str], list[list[object]]]:
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        content = uploaded_file.read()
        for encoding in ("utf-8-sig", "utf-8", "cp950", "latin-1"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValidationError("無法解讀這個 CSV 檔案的編碼。")
        reader = csv.reader(io.StringIO(text))
        rows = [list(row) for row in reader]
    elif filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    else:
        raise ValidationError("目前只支援 .csv 與 .xlsx 檔案。")

    if not rows:
        raise ValidationError("上傳的檔案是空的。")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return headers, rows[1:]


def parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def convert_import_value(field: DataField, raw_value):
    if raw_value in (None, ""):
        return None if field.field_type not in DataField.TEXT_LIKE_TYPES else ""
    try:
        if field.field_type == DataField.TEXT:
            return str(raw_value).strip()
        if field.field_type == DataField.LONG_TEXT:
            return str(raw_value)
        if field.field_type == DataField.INTEGER:
            return int(raw_value)
        if field.field_type == DataField.DECIMAL:
            return Decimal(str(raw_value))
        if field.field_type == DataField.BOOLEAN:
            return parse_bool(raw_value)
        if field.field_type == DataField.DATE:
            if isinstance(raw_value, datetime):
                return raw_value.date()
            if isinstance(raw_value, date):
                return raw_value
            parsed = parse_date(str(raw_value))
            if parsed is None:
                raise ValidationError(f"「{raw_value}」不是欄位 {field.name} 可接受的日期格式。")
            return parsed
        if field.field_type == DataField.DATETIME:
            if isinstance(raw_value, datetime):
                return raw_value
            if isinstance(raw_value, date):
                return datetime.combine(raw_value, time.min)
            parsed = parse_datetime(str(raw_value))
            if parsed is None:
                parsed_date = parse_date(str(raw_value))
                if parsed_date is None:
                    raise ValidationError(f"「{raw_value}」不是欄位 {field.name} 可接受的日期時間格式。")
                return datetime.combine(parsed_date, time.min)
            return parsed
        if field.field_type == DataField.EMAIL:
            return str(raw_value).strip()
        if field.field_type == DataField.URL:
            return str(raw_value).strip()
        if field.field_type == DataField.RELATION:
            if field.related_table is None:
                raise ValidationError(f"欄位 {field.name} 尚未設定關聯資料表。")
            candidate = str(raw_value).strip()
            record = None
            if candidate.isdigit():
                record = field.related_table.records.filter(pk=int(candidate)).first()
            if record is None:
                matching = [
                    item
                    for item in field.related_table.records.all()
                    if item.display_label.strip().lower() == candidate.lower()
                ]
                if len(matching) == 1:
                    record = matching[0]
                elif len(matching) > 1:
                    raise ValidationError(
                        f"關聯值「{raw_value}」在 {field.related_table.name} 中符合多筆記錄。"
                    )
            if record is None:
                raise ValidationError(
                    f"關聯值「{raw_value}」無法在 {field.related_table.name} 中找到對應記錄。"
                )
            return record.pk
        return raw_value
    except ValidationError:
        raise
    except Exception as exc:
        raise ValidationError(f"欄位 {field.name} 的值「{raw_value}」格式不正確。") from exc


@transaction.atomic
def import_rows_to_table(
    table: DataTable,
    headers: list[str],
    rows: list[list[object]],
    *,
    user=None,
    create_missing_fields: bool = False,
) -> dict[str, int]:
    field_map: list[tuple[int, DataField]] = []
    created_fields = 0
    current_fields = list(table.ordered_fields)
    next_order = max([field.order for field in current_fields], default=0) + 10

    for index, header in enumerate(headers, start=1):
        label, slug = normalize_header(header, index)
        field = table.fields.filter(slug=slug).first()
        if field is None:
            field = table.fields.filter(name__iexact=label).first()
        if field is None:
            if not create_missing_fields:
                raise ValidationError(
                    f"欄位「{label}」找不到對應的既有欄位。"
                    "若要匯入，請啟用自動建立缺少欄位。"
                )
            field = DataField.objects.create(
                table=table,
                name=label,
                slug=slug,
                field_type=DataField.TEXT,
                order=next_order,
            )
            next_order += 10
            created_fields += 1
        if user is not None and not field.can_edit(user):
            raise ValidationError(
                f"你沒有權限匯入資料到欄位「{field.name}」。"
            )
        field_map.append((index - 1, field))

    created_records = 0
    for row_number, row in enumerate(rows, start=2):
        if not any(cell not in (None, "") for cell in row):
            continue

        payload = {}
        for field in table.ordered_fields:
            if field.default_value:
                payload[field.slug] = field.serialize_value(
                    field.parse_stored_value(field.default_value)
                )

        for column_index, field in field_map:
            raw_value = row[column_index] if column_index < len(row) else None
            try:
                converted = convert_import_value(field, raw_value)
            except ValidationError as exc:
                raise ValidationError(
                    f"第 {row_number} 列，欄位「{field.name}」：{'；'.join(exc.messages)}"
                ) from exc
            payload[field.slug] = field.serialize_value(converted)

        missing_required = [
            field.name
            for field in table.ordered_fields
            if field.required and payload.get(field.slug) in (None, "")
        ]
        if missing_required:
            joined = ", ".join(missing_required)
            raise ValidationError(
                f"第 {row_number} 列缺少必填欄位：{joined}。"
            )

        DataRecord.objects.create(table=table, data=payload)
        created_records += 1

    return {
        "created_fields": created_fields,
        "created_records": created_records,
    }


def parse_condition_value(field: DataField, raw_value: str):
    raw_value = (raw_value or "").strip()
    if raw_value == "":
        return None
    if field.field_type == DataField.INTEGER:
        return int(raw_value)
    if field.field_type == DataField.DECIMAL:
        return Decimal(raw_value)
    if field.field_type == DataField.BOOLEAN:
        return parse_bool(raw_value)
    if field.field_type == DataField.DATE:
        parsed = parse_date(raw_value)
        if parsed is None:
            raise ValidationError(f"「{raw_value}」不是有效的日期格式。")
        return parsed
    if field.field_type == DataField.DATETIME:
        parsed = parse_datetime(raw_value)
        if parsed is None:
            raise ValidationError(f"「{raw_value}」不是有效的日期時間格式。")
        return parsed
    if field.field_type == DataField.RELATION:
        if raw_value.isdigit():
            return int(raw_value)
        return raw_value.lower()
    return raw_value.lower()


def _normalize_value_for_field(field: DataField, raw_value):
    parsed = field.parse_stored_value(raw_value)
    if field.field_type == DataField.RELATION and raw_value not in (None, ""):
        return {
            "pk": int(raw_value),
            "label": field.display_value(raw_value).lower(),
        }
    if isinstance(parsed, str):
        return parsed.lower()
    return parsed


def _normalize_record_value(record: DataRecord, field_path: FieldPathSpec):
    raw_value = get_value_for_field_path(record, field_path)
    return _normalize_value_for_field(field_path.field, raw_value)


def record_matches_condition(
    record: DataRecord,
    field_path: FieldPathSpec,
    operator: str,
    raw_value: str,
) -> bool:
    current = _normalize_record_value(record, field_path)
    if operator == "is_empty":
        return current in (None, "", {})
    if operator == "not_empty":
        return current not in (None, "", {})
    if operator == "is_true":
        return bool(current) is True
    if operator == "is_false":
        return bool(current) is False

    expected = parse_condition_value(field_path.field, raw_value)

    if field_path.field.field_type == DataField.RELATION and isinstance(current, dict):
        if isinstance(expected, int):
            current_value = current["pk"]
        else:
            current_value = current["label"]
    else:
        current_value = current

    if operator == "equals":
        return current_value == expected
    if operator == "not_equals":
        return current_value != expected
    if operator == "contains":
        return str(expected or "") in str(current_value or "")
    if operator == "starts_with":
        return str(current_value or "").startswith(str(expected or ""))
    if operator == "ends_with":
        return str(current_value or "").endswith(str(expected or ""))
    if current_value in (None, ""):
        return False
    if operator == "gt":
        return current_value > expected
    if operator == "gte":
        return current_value >= expected
    if operator == "lt":
        return current_value < expected
    if operator == "lte":
        return current_value <= expected
    return False


def run_query(records, table: DataTable, conditions: list[dict[str, str]], match_mode: str, *, user=None) -> list[DataRecord]:
    active_conditions = [
        condition
        for condition in conditions
        if (condition.get("field_path") or condition.get("field")) and condition.get("operator")
    ]
    if not active_conditions:
        return list(records)

    field_paths = {}
    if user is None:
        for field in table.ordered_fields:
            field_paths[field.slug] = FieldPathSpec(token=field.slug, label=field.name, field=field)
    else:
        field_paths = {
            spec.token: spec
            for spec in get_accessible_field_paths(table, user, include_joined=True)
        }
    matched = []
    for record in records:
        checks = []
        for condition in active_conditions:
            token = condition.get("field_path") or condition.get("field")
            field_path = field_paths.get(token)
            if field_path is None:
                raise ValidationError(f"欄位路徑「{token}」目前不可用。")
            checks.append(
                record_matches_condition(
                    record,
                    field_path,
                    condition["operator"],
                    condition.get("value", ""),
                )
            )
        if not checks:
            matched.append(record)
        elif match_mode == "any" and any(checks):
            matched.append(record)
        elif match_mode != "any" and all(checks):
            matched.append(record)
    return matched
